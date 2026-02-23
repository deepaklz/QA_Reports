"""
convert_excel.py
Run this script every time you update reports.xlsx to refresh the website data.
Usage:  python convert_excel.py
"""

import openpyxl
import json
import os
import re


EXCEL_FILE = "reports.xlsx"
OUTPUT_JS  = "data.js"


# ── Month parser ─────────────────────────────────────────────────────────────
def parse_month_key(date_range: str) -> str:
    """Extract 'Feb 2026' from '16th - 21th Feb 2026'."""
    months = ["Jan","Feb","Mar","Apr","May","Jun",
              "Jul","Aug","Sep","Oct","Nov","Dec"]
    for m in months:
        match = re.search(rf'\b({m})\s+(\d{{4}})\b', date_range, re.IGNORECASE)
        if match:
            return f"{match.group(1).capitalize()} {match.group(2)}"
    return "Unknown"


# ── Auto-classifier ───────────────────────────────────────────────────────────
def classify_observation(obs: str):
    """
    Rule-based classifier based on the QA taxonomy:
      Error Category : UI | UX | Functionality | Logical/Business Logic | Performance
      Error Impact   : S1 – Critical | S2 – Major | S3 – Minor | S4 – Low/Cosmetic
    """
    t = obs.lower()

    # ── Error Category ────────────────────────────────────────────────────────
    # Performance (check first – highest specificity)
    perf = ["slow", "crash", "timeout", "load time", "memory leak", "latency",
            "100 users", "under load"]

    # Logical / Business Logic
    logic = ["logic", "calculation", "formula", "90%", "100%",
             "wrong count", "chargemix", "incorrect picture",
             "planned chargemix", "liquid metal wt",
             "decimal place", "tapping min", "tapping max",
             "min max value", "should be between"]

    # UX – how it works / feels (but still functional)
    ux = ["should not show", "better to", "counter-intuitive", "confusing",
          "too many click", "not similar", "format are not",
          "validation.*remark", "error remark.*display", "clarity",
          "filter before", "casting type filter", "before grade",
          "repeated grade", "fix size", "default.*size",
          "expand but do not", "no workaround",
          "every validation"]

    # UI – visual / cosmetic
    ui = ["fade", "spelling", "typo", "blurry", "misalign", "overlap",
          "color", "colour", "font", "icon", "logo", "pixel",
          "button.*size", "box size", "format.*date", "date.*format",
          "whatsapp", "whatapp", "spacing", "visual"]

    if any(k in t for k in perf):
        category = "Performance"
    elif any(re.search(k, t) for k in logic):
        category = "Logical/Business Logic"
    elif any(re.search(k, t) for k in ux):
        category = "UX"
    elif any(re.search(k, t) for k in ui):
        category = "UI"
    else:
        category = "Functionality"

    # ── Error Impact ──────────────────────────────────────────────────────────
    # S1 – Critical / Blocker
    s1 = ["crash", "cannot login", "app.*crash", "system.*unusable", "blocker"]

    # S2 – Major (primary function broken)
    s2 = ["failed to create", "failed to", "error not a valid", "not a valid json",
          "wrong count", "incorrect picture", "please retry",
          "tapping min max", "error occured", "error.*remark.*not show",
          "logic recheck", "90%"]

    # S4 – Low / Cosmetic
    s4 = ["fade", "spelling", "typo", "whatapp", "whatsapp",
          "2 pixels", "cosmetic", "blurry", "should not show.*-",
          "better to use", r"use \"|\"",
          "decimal place", "format.*not similar", "not similar.*format",
          "box size.*not fix", "repeated grade"]

    if any(re.search(k, t) for k in s1):
        impact = "S1 – Critical"
    elif any(re.search(k, t) for k in s2):
        impact = "S2 – Major"
    elif any(re.search(k, t) for k in s4):
        impact = "S4 – Low/Cosmetic"
    else:
        impact = "S3 – Minor"

    return category, impact


# ── Excel → JSON ──────────────────────────────────────────────────────────────
def excel_to_json(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    # Find index of the "Links" column
    links_col_idx = next(
        (i for i, h in enumerate(headers) if h and h.strip().lower() == "links"),
        None
    )
    # Find indexes of the new classification columns (may or may not exist yet)
    cat_col_idx = next(
        (i for i, h in enumerate(headers) if h and "category" in h.lower()),
        None
    )
    imp_col_idx = next(
        (i for i, h in enumerate(headers) if h and "impact" in h.lower()),
        None
    )

    data = []
    auto_filled = 0

    for row in ws.iter_rows(min_row=2):
        values = [cell.value for cell in row]
        if not any(v is not None and str(v).strip() != "" for v in values):
            continue

        row_dict = {}
        for i, cell in enumerate(row):
            if i < len(headers) and headers[i]:
                header = headers[i]
                value  = str(cell.value).strip() if cell.value is not None else ""

                # Extract real hyperlink from Links column
                if i == links_col_idx and cell.hyperlink and cell.hyperlink.target:
                    row_dict[header] = str(cell.hyperlink.target).strip()
                else:
                    row_dict[header] = value

        # Auto-classify if columns are missing or empty in Excel
        obs = row_dict.get("Observations", "")
        existing_cat = (row_dict.get("Error Category") or "").strip()
        existing_imp = (row_dict.get("Error Impact")   or "").strip()

        if obs and (not existing_cat or not existing_imp):
            auto_cat, auto_imp = classify_observation(obs)
            if not existing_cat:
                row_dict["Error Category"] = auto_cat
            if not existing_imp:
                row_dict["Error Impact"] = auto_imp
            auto_filled += 1

        # Attach computed month label
        row_dict["_month"] = parse_month_key(row_dict.get("Date Range", ""))
        data.append(row_dict)

    return headers, data, auto_filled


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    if not os.path.exists(EXCEL_FILE):
        print(f"ERROR: '{EXCEL_FILE}' not found in current directory.")
        return

    headers, data, auto_filled = excel_to_json(EXCEL_FILE)

    js_content = f"""// AUTO-GENERATED by convert_excel.py — do not edit manually.
// Run 'python convert_excel.py' after updating reports.xlsx

const REPORT_HEADERS = {json.dumps(headers, ensure_ascii=False, indent=2)};

const REPORT_DATA = {json.dumps(data, ensure_ascii=False, indent=2)};
"""
    with open(OUTPUT_JS, "w", encoding="utf-8") as f:
        f.write(js_content)

    print(f"✅ Done! Exported {len(data)} rows to {OUTPUT_JS}")
    if auto_filled:
        print(f"   🤖 Auto-classified {auto_filled} rows (Error Category + Error Impact)")
    months = sorted(set(r["_month"] for r in data))
    print(f"   Months found: {', '.join(months)}")

    # Preview classification
    print()
    print(f"{'#':<4} {'Category':<25} {'Impact':<18} Observation[:50]")
    print("─" * 90)
    for r in data:
        sn  = r.get("Sn", "")
        cat = r.get("Error Category", "—")
        imp = r.get("Error Impact", "—")
        obs = (r.get("Observations", "") or "")[:50].replace("\n", " ")
        print(f"{sn:<4} {cat:<25} {imp:<18} {obs}")


if __name__ == "__main__":
    main()
